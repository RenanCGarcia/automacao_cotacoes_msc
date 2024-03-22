from selenium import webdriver
import datetime
import openpyxl
import os
from time import sleep

# Abre o chrome no site de cotações da MSC
driver = webdriver.Chrome()
driver.get("https://mscmkt3.com/")

# Coleta as informações a serem colocadas na planilha
data = datetime.date.today()
cabecalho = driver.find_element("xpath", "/html/body/div/table/tbody/tr[1]/td").text
dolar = driver.find_element("xpath", "/html/body/div/table/tbody/tr[3]/td[2]").text
euro = driver.find_element("xpath", "/html/body/div/table/tbody/tr[5]/td[2]").text
libra = driver.find_element("xpath", "/html/body/div/table/tbody/tr[7]/td[2]").text

# Entra na planilha(procura a planilha do dia anterior, caso não ache, procura a última criada em até 1 mês)
diretorio = os.path.dirname(__file__)
for dia in range(1,32):
    arquivo = os.path.join(diretorio,f"{data - datetime.timedelta(days=dia)}_cotacao_MSC.xlsx")
    print(arquivo)
    try:
        if os.path.exists(arquivo):
            print(f"Arquivo existe em: {arquivo}")
            planilha = openpyxl.load_workbook(arquivo)
            break
    except NameError:
        sleep(10)

# Seleciona página
cotacao = planilha['cotação']

# Adiciona as informações na planilha
cotacao.append([cabecalho,data,dolar,euro,libra])

# Salva planilha com a data de hoje
planilha.save(f'{diretorio}/{data}_cotacao_MSC.xlsx')

#DEBUG
print(f"""
Cabeçalho: {cabecalho}
Data: {data}
Dolar: {dolar}
Euro: {euro}
Libra: {libra}
""")

# Como visualizar páginas existentes
# print(planilha.sheetnames)

# Como criar uma página
#planilha.create_sheet('cotação')